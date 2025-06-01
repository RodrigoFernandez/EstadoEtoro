"""
procesadores.py
"""
from openpyxl import load_workbook
import pandas as pd
from etoro_modelos.modelos import Deposito, DetalleTenencia, Tenencia
from etoro_modelos.modelos import DividendoMensual, Etoro
from .constantes import Columnas
from .constantes import ColumnasReporteEtoro as ColRepE


class Generador(object):
    def __init__(self, ruta):
        self.ruta = ruta

    def procesar_fila(self, fila):
        unidades = fila[ColRepE.UNIDADES.value].value
        if fila[ColRepE.UNIDADES.value].value == Columnas.DEPOSITO:
            unidades = ""

        id_de_poscion = ""
        if fila[ColRepE.ID_POSICION.value].value is not None:
            id_de_poscion = fila[ColRepE.ID_POSICION.value].value

        rta = {
            Columnas.FECHA: fila[ColRepE.FECHA.value].value,
            Columnas.TIPO: fila[ColRepE.TIPO.value].value,
            Columnas.DETALLES: fila[ColRepE.DETALLES.value].value,
            Columnas.IMPORTE: fila[ColRepE.IMPORTE.value].value,
            Columnas.UNIDADES: unidades,
            Columnas.ID_DE_POSICION: id_de_poscion,
        }

        return rta

    def procesar_datos(self, datos):
        df = pd.DataFrame(datos)
        df.set_index(Columnas.FECHA, inplace=True)
        agrupados = df.groupby(Columnas.ID_DE_POSICION)

        depositos = self.convert_2_depositos(agrupados.get_group(Columnas.DEPOSITO))
        tenencias, comisiones = self.convert_2_tenencias(agrupados)

        return depositos, tenencias, comisiones

    def convert_2_depositos(self, depos):
        rta = [Deposito(dep[Columnas.IMPORTE],
                        dep[Columnas.DETALLES],
                        indice_fila_fecha)
               for indice_fila_fecha, dep in depos.iterrows()]
        return rta

    def get_ticket(self, detalle):
        if detalle:
            partes = detalle.split("/")
            return partes[0]
        return ""

    def convert_2_tenencias(self, tenencias):
        rta = []
        comisiones = []
        for clave in tenencias.groups.keys():
            if clave not in ['', Columnas.DEPOSITO]:
                una_tenencia = tenencias.get_group(clave)

                es_primer_movimiento = True
                ticket = ""
                detalles = []
                for indice_fecha, t in una_tenencia.iterrows():
                    if t[Columnas.TIPO] != 'Comisión':
                        if es_primer_movimiento:
                            es_primer_movimiento = False
                            ticket = self.get_ticket(t[Columnas.DETALLES])
                        detalles.append(DetalleTenencia(
                            t[Columnas.IMPORTE],
                            indice_fecha,
                            t[Columnas.DETALLES],
                            t[Columnas.UNIDADES]))
                    else:
                        comisiones.append(DetalleTenencia(
                            t[Columnas.IMPORTE],
                            indice_fecha,
                            t[Columnas.DETALLES],
                            t[Columnas.UNIDADES]))
                rta.append(Tenencia(clave, ticket, detalles))

        a = sorted(rta, key=lambda t: t.ticket)
        return (a, comisiones)

    def preprocesar_actividad(self, hoja):
        es_titulo = True
        pre_procesadas = []

        for fila in hoja.rows:
            if es_titulo:
                es_titulo = False
            else:
                pre_procesadas.append(self.procesar_fila(fila))
        return pre_procesadas

    def convert_2_dividendos_mensuales(self, sumatorias):
        rta = []

        for index, fila in sumatorias.iterrows():
            rta.append(DividendoMensual(fila[Columnas.ANIO],
                       fila[Columnas.MES], fila[Columnas.IMPORTE]))

        return rta

    def procesar_fila_div(self, fila):
        rta = {
            Columnas.FECHA: fila[0].value,
            Columnas.DETALLES: fila[1].value,
            Columnas.IMPORTE: fila[2].value,
            Columnas.ID_DE_POSICION: fila[5].value,
        }

        return rta

    def preprocesar_dividendos(self, hoja):
        es_titulo = True
        pre_procesadas = []

        for fila in hoja.rows:
            if es_titulo:
                es_titulo = False
            else:
                pre_procesadas.append(self.procesar_fila_div(fila))
        return pre_procesadas

    def procesar_datos_dividendos(self, datos):
        df = pd.DataFrame(datos)

        df[Columnas.FECHA2] = pd.to_datetime(df[Columnas.FECHA],
                                             format="%d/%m/%Y")

        df[Columnas.ANIO] = pd.DatetimeIndex(df[Columnas.FECHA2]).year
        df[Columnas.MES] = pd.DatetimeIndex(df[Columnas.FECHA2]).month

        sumatorias = df.groupby([Columnas.ANIO, Columnas.MES]).agg(
            {Columnas.IMPORTE: 'sum'}).reset_index()
        sumatorias.sort_values(by=[Columnas.ANIO, Columnas.MES],
                               inplace=True,
                               ascending=False)
        rta = self.convert_2_dividendos_mensuales(sumatorias)
        return rta

    def procesar_fila_cerrada(self, fila):
        rta = {
            Columnas.ID_DE_POSICION: fila[0].value,
        }

        return rta

    def preprocesar_cerradas(self, hoja):
        es_titulo = True
        pre_procesadas = []

        for fila in hoja.rows:
            if es_titulo:
                es_titulo = False
            else:
                pre_procesadas.append(self.procesar_fila_cerrada(fila))
        return pre_procesadas

    def procesar_cerradas(self, datos):
        aux = []
        for cierre in datos:
            aux.append(cierre[Columnas.ID_DE_POSICION])

        aux = set(aux)
        return aux

    def leer_excel(self):
        planilla = load_workbook(self.ruta)
        hoja_actividades = planilla[planilla.sheetnames[2]]

        pre_procesadas = self.preprocesar_actividad(hoja_actividades)

        depositos, posiciones, comisiones = self.procesar_datos(pre_procesadas)

        hoja_dividendos = planilla[planilla.sheetnames[3]]
        pre_procesadas_div = self.preprocesar_dividendos(hoja_dividendos)
        dividendos = self.procesar_datos_dividendos(pre_procesadas_div)

        hoja_cerradas = planilla[planilla.sheetnames[1]]
        pre_procesadas_cerradas = self.preprocesar_cerradas(hoja_cerradas)
        cerradas = self.procesar_cerradas(pre_procesadas_cerradas)

        return depositos, posiciones, dividendos, cerradas, comisiones

    def generar_etoro(self):
        depositos, posiciones, dividendos, cerradas, comisiones = self.leer_excel()
        rta = Etoro(depositos, posiciones, dividendos, cerradas, comisiones)
        return rta
